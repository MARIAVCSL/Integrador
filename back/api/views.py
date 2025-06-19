from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from .models import Sensores, Historico, Ambientes
from rest_framework import generics
from django.contrib.auth.models import User
from .serializers import SensoresSerializer, HistoricoSerializer, AmbientesSerializer,  UserSerializer
from openpyxl import load_workbook, Workbook
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from django.utils.dateparse import parse_date
from datetime import datetime

class SignUpView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

# ----------- Ambientes -------------

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_ambientes(request):
    if request.method == 'GET':
        ambientes = Ambientes.objects.all()
        serializer = AmbientesSerializer(ambientes, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = AmbientesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AmbientesView(ListCreateAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['sig']
    search_fields = ['sig']

class AmbientesDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]

# ----------- Sensores -------------

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_sensores(request):
    if request.method == 'GET':
        sensores = Sensores.objects.all()
        serializer = SensoresSerializer(sensores, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = SensoresSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SensoresView(ListCreateAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'sensor', 'status']
    search_fields = ['id', 'sensor', 'status']

class SensoresDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    permission_classes = [IsAuthenticated]


# ----------- Histórico -------------

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_historico(request):
    if request.method == 'GET':
        historico = Historico.objects.all()
        serializer = HistoricoSerializer(historico, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = HistoricoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class HistoricoView(ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'timestamp']
    search_fields = ['id', 'timestamp']

    def get_queryset(self):
        queryset = super().get_queryset()
        sensor_id = self.request.query_params.get('sensor')
        hora_str = self.request.query_params.get('hora') 
        data_str = self.request.query_params.get('data')

        if sensor_id:
            queryset = queryset.filter(sensor__id=sensor_id)

        if data_str:
            queryset = queryset.filter(timestamp__date=data_str)

        if hora_str:
            try:
                hora_int = int(hora_str)
                queryset = queryset.filter(timestamp__hour=hora_int)            
            except ValueError:
                pass  

        return queryset
    

class HistoricoDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]


class UploadXLSXView_amb(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  
            sig = str(row[0]).strip() if row[0] is not None else None
            descricao = str(row[1]).strip() if row[1] is not None else None
            ni = str(row[2]).strip() if row[2] is not None else None
            responsavel = str(row[3]).strip() if row[3] is not None else None
            
            if not ni:
                    print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                    continue  
                
            Ambientes.objects.create(
                sig=row[0],
                descricao=row[1],
                ni=row[2],
                responsavel=row[3],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})

class UploadXLSXView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  
            sensor = str(row[0]).strip() if row[0] is not None else None
            mac_addres = str(row[1]).strip() if row[1] is not None else None
            unidade_med = str(row[2]).strip() if row[2] is not None else None
            latitude = str(row[3]).strip() if row[3] is not None else None
            longitude = str(row[4]).strip() if row[4] is not None else None
            status = str(row[5]).strip() if row[5] is not None else None
            
            if not sensor:
                    print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                    continue  
                
            Sensores.objects.create(
                sensor=row[0],
                mac_addres=row[1],
                unidade_med=row[2],
                latitude=row[3],
                longitude=row[4],
                status=row[5],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})
    
class ImportHistoricoView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            valor = str(row[0]).strip() if row[0] is not None else None
            timestamp = row[1]
            ambiente = row[2]
            sensor = row[3]

            try:
                sensor = Sensores.objects.get(id=sensor)
            except Sensores.DoesNotExist:
                print(f"[Linha {i+2}] Sensor com ID {sensor} não encontrado.")
                continue

            try:
                ambiente = Ambientes.objects.get(id=ambiente)
            except Ambientes.DoesNotExist:
                print(f"[Linha {i+2}] Ambiente com ID {ambiente} não encontrado..")
                continue

            Historico.objects.create(
                valor=valor,
                timestamp=timestamp,
                ambiente=ambiente,
                sensor=sensor,
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})

@api_view(['GET'])
def exportar_xlsx_ambientes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    # Cabeçalhos
    headers = ['sig', 'descricao', 'ni', 'responsavel']
    ws.append(headers)

    # Dados das ordens
    for ambientes in Ambientes.objects.all():
        ws.append([
            ambientes.sig,
            ambientes.descricao,
            ambientes.ni,
            ambientes.responsavel,
        ])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    # Gerar resposta HTTP com arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ordens_servico.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
def exportar_xlsx_sensores(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensores"

    # Cabeçalhos
    headers = ['sensor', 'mac_addres', 'unidade_med', 'latitude', 'longitude', 'status']
    ws.append(headers)

    # Dados dos dos sensores dens
    for sensores in Sensores.objects.all():
        ws.append([
            sensores.sensor,
            sensores.mac_addres,
            sensores.unidade_med,
            sensores.latitude,
            sensores.longitude,
            sensores.status,
        ])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    # Gerar resposta HTTP com arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ordens_servico.xlsx"'
    wb.save(response)
    return response

# @api_view(['GET'])
# def exportar_xlsx_historico(request):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Historico"

#     # Cabeçalhos
#     headers = ['Sensor', 'Ambiente', 'Valor', 'Timestamp']
#     ws.append(headers)

#     # Dados do histórico
#     for historico in Historico.objects.select_related('sensor', 'ambiente').all():
#         ws.append([
#             str(historico.sensor),
#             str(historico.ambiente),
#             historico.valor,
#             historico.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
#         ])

#     # Ajustar largura das colunas
#     for col in ws.columns:
#         max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#         ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

#     # Gerar resposta HTTP com o arquivo
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'
#     wb.save(response)
#     return response

@api_view(['GET'])
def exportar_xlsx_historico(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    # Cabeçalhos
    headers = ['valor', 'timestamp', 'ambiente_id', 'sensor_id']
    ws.append(headers)

    for hist in Historico.objects.all():
        ws.append([
            hist.valor,
            hist.timestamp.strftime('%Y-%m-%d %H:%M:%S') if hist.timestamp else '',
            hist.ambiente.id if hist.ambiente else '',
            hist.sensor.id if hist.sensor else '',
        ])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'
    wb.save(response)
    return response


def exportar_historico_excel(request):
    hora = request.GET.get('hora')
    data = request.GET.get('data')

    if not hora or not data:
        return HttpResponse("Parâmetros 'hora' e 'data' são obrigatórios.", status=400)

    try:
        data_obj = datetime.strptime(data, "%Y-%m-%d").date()
        hora_int = int(hora)

        historico = Historico.objects.filter(
            timestamp__date=data_obj,
            timestamp__hour=hora_int
        )

    except ValueError:
        return HttpResponse("Formato inválido para data ou hora.", status=400)

    # Criação do arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    # Cabeçalhos
    headers = ['ID', 'Sensor', 'Ambiente', 'Valor', 'Timestamp']
    ws.append(headers)

    # Dados
    for h in historico:
        ws.append([
            h.id,
            str(h.sensor),
            str(h.ambiente),
            h.valor,
            h.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Resposta HTTP com o Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=historico_{data}_{hora}h.xlsx'
    wb.save(response)
    return response